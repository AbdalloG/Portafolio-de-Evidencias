from unittest import result
from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    resultado = hunter.domain_search(company=organizacion, \
        limit=5, emails_type='personal')
    return resultado
def clean_data():
    clean_list = []
    for email in datosEncontrados['emails']:
        value_dict = {}
        value_dict['email'] = email['value']
        clean_list.append(value_dict)
    return clean_list
def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja["A1"] = "Emails"
    count = 2
    for email in datosEncontrados['emails']:
        hoja.cell(count, 1, email['value'])
        count = count + 1
    libro.save("Hunter" + organizacion + ".xlsx")

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados == None:
    exit()
else:
    print(datosEncontrados)
    GuardarInformacion(datosEncontrados, orga)